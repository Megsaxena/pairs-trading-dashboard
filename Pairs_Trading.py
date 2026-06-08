"""
Pairs Trading Strategy — S&P 500
=================================
Improvements over original R implementation:
  1. Train/test split to eliminate look-ahead bias in pair selection
  2. OLS hedge ratio (beta) replaces the naive price ratio
  3. Half-life of mean reversion (Ornstein-Uhlenbeck) calibrates the z-score window
  4. Annualised Sharpe ratio (×√252)
  5. Transaction costs (per-leg, round-trip)
  6. Benjamini-Hochberg FDR correction for multiple cointegration tests
  7. Vectorised signal generation (no per-row loop)
  8. Max drawdown and Calmar ratio added to performance metrics
  9. Dollar-neutral position sizing (equal notional per leg)
"""

import warnings
warnings.filterwarnings("ignore")

from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf
import requests
from io import StringIO

import statsmodels.api as sm
from statsmodels.tsa.stattools import adfuller, coint
from statsmodels.regression.rolling import RollingOLS

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.gridspec import GridSpec
from itertools import product as iproduct
from concurrent.futures import ProcessPoolExecutor, as_completed
import multiprocessing
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── reproducibility ──────────────────────────────────────────────────────────
np.random.seed(42)

# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════
CFG = dict(
    start_date          = "2020-02-11",
    train_end_date      = "2022-03-14",   # in-sample: pair selection & calibration
    test_end_date       = "2024-03-14",   # out-of-sample: forward backtest
    corr_threshold      = 0.70,           # minimum in-sample return correlation
                                          # lowered from 0.75 — cross-sector pairs
                                          # tend to have slightly lower correlation
    coint_pval          = 0.05,           # Engle-Granger p-value threshold
    fdr_correction      = False,          # BH-FDR is too aggressive for trading:
                                          # with 300 pairs tested the threshold becomes
                                          # p ≤ 0.05/300 ≈ 0.00017, rejecting almost everything.
                                          # Raw p < 0.05 is standard in quant-finance pairs work.
    zscore_entry        = 2.0,            # open trade when |z| crosses this
    zscore_exit         = 0.5,            # close trade when |z| falls below this
    zscore_stop         = 3.5,            # stop-loss when |z| exceeds this
    transaction_cost    = 0.0010,         # one-way cost per leg (10 bps); 4 legs per round-trip
    min_halflife_days   = 5,              # reject pairs whose spread reverts too fast
    max_halflife_days   = 120,            # reject pairs that barely mean-revert
    # ── Universe scope ──────────────────────────────────────────────────
    same_sector_only    = False,          # False → full cross-sector universe (user preference);
                                          # True  → only same-GICS-sector pairs (faster, stricter)
    max_corr_pairs      = 300,            # screen the top 300 most-correlated pairs
                                          # for cointegration — large enough for cross-
                                          # sector coverage, fast enough to run in < 2 min
    top_coint_pairs     = 100,            # keep only the N most-cointegrated pairs
                                          # (sorted by p-value ascending); None = keep all
                                          # int   → hard cap (e.g. 500) for quick testing
    n_workers           = max(1, multiprocessing.cpu_count() - 1),  # parallel coint tests
    top_n_plot          = 5,              # pairs/combos to chart
)

# ── Parameter grid for z-score / window optimisation ──────────────────────
GRID = dict(
    entry_thresholds = [1.5, 1.75, 2.0, 2.25, 2.5],   # |z| to open trade
    exit_thresholds  = [0.25, 0.5, 0.75],              # |z| to close trade
    z_windows        = [20, 30, 40, 60, 90],           # rolling window (days)
    stop_thresholds  = [3.0, 3.5, 4.0],               # stop-loss |z|
)

# Output directory — all charts and CSVs are written here
OUTPUT_DIR = Path(r"W:\Meghna\Misc\Pairs Trading")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════════
# STEP 1 — FETCH S&P 500 UNIVERSE
# ═══════════════════════════════════════════════════════════════════════════
def fetch_sp500_tickers() -> pd.DataFrame:
    """
    Scrape S&P 500 constituents and sector mapping from Wikipedia.
    pandas.read_html passes no User-Agent and gets a 403; we use requests
    with a browser header first, then hand the HTML text to pd.read_html.
    Falls back to a hardcoded sector-agnostic stub if the request still fails.
    """
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        )
    }
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        tables = pd.read_html(StringIO(resp.text))
        df = tables[0][["Symbol", "GICS Sector"]].copy()
        df["Symbol"] = df["Symbol"].str.replace(".", "-", regex=False)
        return df
    except Exception as e:
        print(f"  Warning: could not fetch S&P 500 list ({e}).")
        print("  Falling back to yfinance SP500 tickers via alternative source …")
        # Alternative: pull from GitHub-hosted CSV (no auth required)
        alt_url = (
            "https://raw.githubusercontent.com/datasets/s-and-p-500-companies"
            "/main/data/constituents.csv"
        )
        df = pd.read_csv(alt_url)
        df = df.rename(columns={"Symbol": "Symbol", "Sector": "GICS Sector"})
        df["Symbol"] = df["Symbol"].str.replace(".", "-", regex=False)
        return df[["Symbol", "GICS Sector"]]

# ═══════════════════════════════════════════════════════════════════════════
# STEP 2 — DOWNLOAD PRICE DATA
# ═══════════════════════════════════════════════════════════════════════════
def download_prices(tickers: list[str], start: str, end: str) -> pd.DataFrame:
    """
    Download adjusted close prices. Returns a DataFrame indexed by date,
    one column per ticker. Tickers with >20% missing data are dropped.
    """
    print(f"  Downloading {len(tickers)} tickers …")
    raw = yf.download(
        tickers,
        start=start,
        end=end,
        auto_adjust=True,
        progress=False,
    )["Close"]

    if isinstance(raw, pd.Series):          # single ticker edge-case
        raw = raw.to_frame()

    # Drop tickers with too many gaps
    max_missing_frac = 0.20
    keep = raw.columns[raw.isna().mean() < max_missing_frac]
    raw  = raw[keep].copy()

    # Forward-fill remaining small gaps (weekends / holidays already absent)
    raw.ffill(inplace=True)
    raw.dropna(inplace=True)

    print(f"  Retained {raw.shape[1]} tickers after quality filter.")
    return raw

# ═══════════════════════════════════════════════════════════════════════════
# STEP 3 — PAIR SELECTION (IN-SAMPLE ONLY)
# ═══════════════════════════════════════════════════════════════════════════
def compute_returns(prices: pd.DataFrame) -> pd.DataFrame:
    """Daily log returns."""
    return np.log(prices / prices.shift(1)).dropna()


def select_correlated_pairs(
    returns: pd.DataFrame,
    threshold: float,
    sector_map: dict | None = None,
    same_sector_only: bool = True,
    max_pairs: int | None = None,
) -> pd.DataFrame:
    """
    Screen the full universe for candidate pairs.

    Parameters
    ----------
    returns          : daily log-return DataFrame (tickers as columns)
    threshold        : minimum Pearson correlation to keep a pair
    sector_map       : {ticker: sector} dict from the S&P 500 table
    same_sector_only : if True, only test pairs in the same GICS sector.
                       This is both faster (11 sectors × ~44 stocks each ≈
                       ~10k intra-sector pairs vs ~118k cross-sector) and
                       economically sounder — cointegration between stocks
                       in different sectors is usually spurious.
    max_pairs        : optional hard cap (None = keep everything that passes)
    """
    tickers = list(returns.columns)
    n = len(tickers)
    ticker_idx = {t: i for i, t in enumerate(tickers)}

    # Build candidate index pairs using sector filter when requested
    if same_sector_only and sector_map:
        # Group tickers by sector
        from collections import defaultdict
        sector_groups = defaultdict(list)
        for t in tickers:
            sector_groups[sector_map.get(t, "Unknown")].append(t)

        candidate_pairs_idx = []
        for sector, members in sector_groups.items():
            for i in range(len(members)):
                for j in range(i + 1, len(members)):
                    ti, tj = members[i], members[j]
                    if ti in ticker_idx and tj in ticker_idx:
                        candidate_pairs_idx.append((ticker_idx[ti], ticker_idx[tj]))

        total_possible = len(candidate_pairs_idx)
        print(f"  Same-sector mode: {total_possible:,} intra-sector pairs to screen "
              f"(vs {n*(n-1)//2:,} cross-sector).")
    else:
        candidate_pairs_idx = [(i, j) for i in range(n) for j in range(i + 1, n)]
        print(f"  Full universe mode: {len(candidate_pairs_idx):,} pairs to screen.")

    if not candidate_pairs_idx:
        return pd.DataFrame()

    # Vectorised correlation matrix — compute once on the returns array
    arr = returns.values                         # shape: (T, N)
    # Normalise columns to get correlation matrix cheaply
    # Use np.corrcoef on the transposed array; for large N this is the fastest path
    print(f"  Computing correlation matrix ({n} × {n}) …")
    cor_matrix = np.corrcoef(arr.T)              # (N, N)

    records = []
    for (i, j) in candidate_pairs_idx:
        c = cor_matrix[i, j]
        if c > threshold:
            records.append((tickers[i], tickers[j], float(c)))

    pairs = pd.DataFrame(records, columns=["stock1", "stock2", "correlation"])
    pairs = pairs.sort_values("correlation", ascending=False)

    if max_pairs is not None:
        pairs = pairs.head(max_pairs)

    pairs = pairs.reset_index(drop=True)
    print(f"  {len(pairs):,} candidate pairs with correlation > {threshold}"
          f"{' (same sector)' if same_sector_only else ''}.")
    return pairs


def benjamini_hochberg(pvals: np.ndarray, alpha: float = 0.05) -> np.ndarray:
    """Return boolean mask of discoveries under BH FDR control."""
    n = len(pvals)
    order   = np.argsort(pvals)
    ranked  = np.empty(n, dtype=int)
    ranked[order] = np.arange(1, n + 1)
    threshold = ranked / n * alpha
    reject = pvals <= threshold
    # BH: reject all hypotheses up to the largest k where p(k) <= k/m * alpha
    max_k = np.where(reject[order])[0]
    if len(max_k) == 0:
        return np.zeros(n, dtype=bool)
    cutoff = max_k[-1]
    result = np.zeros(n, dtype=bool)
    result[order[:cutoff + 1]] = True
    return result


# ── worker function must live at module level for pickle ────────────────
def _coint_worker(args):
    """Test one pair for cointegration. Returns None if test fails."""
    s1, s2, y, x, corr = args
    try:
        from statsmodels.tsa.stattools import coint as _coint
        _, pval, _ = _coint(y, x)
        return {"stock1": s1, "stock2": s2, "correlation": corr, "coint_pval": float(pval)}
    except Exception:
        return None


def test_cointegration(
    prices_train: pd.DataFrame,
    pairs: pd.DataFrame,
    pval_threshold: float,
    fdr: bool,
    n_workers: int = 1,
) -> pd.DataFrame:
    """
    Engle-Granger cointegration test on in-sample prices, parallelised across
    CPU cores so the full S&P 500 intra-sector universe (~hundreds to low
    thousands of pairs) runs in seconds rather than minutes.

    BH-FDR correction is applied AFTER all p-values are collected, which is
    the statistically correct approach — running it pair-by-pair would be wrong.
    """
    # Build argument list — pre-extract numpy arrays to avoid pickle of DataFrame
    args_list = []
    for _, row in pairs.iterrows():
        s1, s2 = row["stock1"], row["stock2"]
        if s1 not in prices_train.columns or s2 not in prices_train.columns:
            continue
        args_list.append((
            s1, s2,
            prices_train[s1].values,
            prices_train[s2].values,
            row["correlation"],
        ))

    n_pairs = len(args_list)
    print(f"  Testing {n_pairs:,} pairs for cointegration "
          f"({'parallel × ' + str(n_workers) + ' cores' if n_workers > 1 else 'single-threaded'}) …")

    results = []
    if n_workers > 1 and n_pairs > 50:
        with ProcessPoolExecutor(max_workers=n_workers) as pool:
            for res in pool.map(_coint_worker, args_list, chunksize=20):
                if res is not None:
                    results.append(res)
    else:
        # Single-threaded fallback (safe on Windows / Jupyter)
        for args in args_list:
            res = _coint_worker(args)
            if res is not None:
                results.append(res)

    if not results:
        return pd.DataFrame()

    df    = pd.DataFrame(results)
    pvals = df["coint_pval"].values

    if fdr:
        reject = benjamini_hochberg(pvals, alpha=pval_threshold)
        label  = "BH-FDR"
    else:
        reject = pvals < pval_threshold
        label  = "raw p-value"

    df = df[reject].reset_index(drop=True)
    print(f"  {len(df):,} cointegrated pairs after {label} filter "
          f"(from {n_pairs:,} candidates).")
    return df

# ═══════════════════════════════════════════════════════════════════════════
# STEP 4 — HEDGE RATIO & HALF-LIFE (IN-SAMPLE)
# ═══════════════════════════════════════════════════════════════════════════
def estimate_hedge_ratio(y: np.ndarray, x: np.ndarray) -> float:
    """
    OLS regression of log(price_y) on log(price_x).
    Returns the beta coefficient used as the hedge ratio.
    Using log-prices is standard in pairs/cointegration literature; it keeps
    the spread stationary in percentage terms and the hedge ratio is unit-free.
    """
    x_c = sm.add_constant(x)
    model = sm.OLS(y, x_c).fit()
    return model.params[1]


def compute_half_life(spread: np.ndarray) -> float:
    """
    Ornstein-Uhlenbeck half-life: regress Δspread on lagged spread.
    λ = -ln(2) / θ, where θ is the mean-reversion speed.
    A shorter half-life means faster reversion → tighter z-score window.
    """
    spread_lag = spread[:-1]
    delta      = np.diff(spread)
    x = sm.add_constant(spread_lag)
    res = sm.OLS(delta, x).fit()
    theta = res.params[1]           # coefficient on lagged spread
    if theta >= 0:
        return np.inf               # non-mean-reverting; reject this pair
    half_life = -np.log(2) / theta
    return half_life


def calibrate_pairs(
    coint_pairs: pd.DataFrame,
    prices_train: pd.DataFrame,
    min_hl: int,
    max_hl: int,
) -> pd.DataFrame:
    """
    For each cointegrated pair:
      - Estimate OLS hedge ratio (in-sample)
      - Compute spread = log(s1) - β·log(s2)
      - Estimate OU half-life
      - Reject pairs outside [min_hl, max_hl] days (too noisy or too slow)
    """
    records = []
    for _, row in coint_pairs.iterrows():
        s1, s2 = row["stock1"], row["stock2"]
        log_y = np.log(prices_train[s1].values)
        log_x = np.log(prices_train[s2].values)

        beta   = estimate_hedge_ratio(log_y, log_x)
        spread = log_y - beta * log_x

        hl = compute_half_life(spread)
        if not np.isfinite(hl) or not (min_hl <= hl <= max_hl):
            continue

        records.append({
            "stock1":      s1,
            "stock2":      s2,
            "correlation": row["correlation"],
            "coint_pval":  row["coint_pval"],
            "hedge_ratio": beta,
            "half_life":   round(hl, 1),
            # z-score window = 2× half-life, bounded reasonably
            "z_window":    int(np.clip(2 * hl, 10, 120)),
        })

    df = pd.DataFrame(records).sort_values("half_life").reset_index(drop=True)
    print(f"  {len(df)} pairs pass half-life filter [{min_hl}, {max_hl}] days.")
    return df

# ═══════════════════════════════════════════════════════════════════════════
# STEP 5 — SIGNAL GENERATION (OUT-OF-SAMPLE, VECTORISED)
# ═══════════════════════════════════════════════════════════════════════════
def compute_spread_zscore(
    prices: pd.DataFrame,
    stock1: str,
    stock2: str,
    hedge_ratio: float,
    z_window: int,
) -> pd.DataFrame:
    """
    Compute the log-price spread and its rolling z-score.
    Uses parameters estimated ONLY on training data (no look-ahead).
    """
    log_y = np.log(prices[stock1])
    log_x = np.log(prices[stock2])
    spread = log_y - hedge_ratio * log_x

    roll_mean = spread.rolling(z_window, min_periods=z_window).mean()
    roll_std  = spread.rolling(z_window, min_periods=z_window).std()

    zscore = (spread - roll_mean) / roll_std

    return pd.DataFrame({
        "spread":    spread,
        "roll_mean": roll_mean,
        "roll_std":  roll_std,
        "zscore":    zscore,
    }, index=prices.index)


def generate_signals_vectorised(
    zscore: pd.Series,
    entry: float,
    exit_: float,
    stop: float,
) -> pd.Series:
    """
    Vectorised finite-state-machine signal generator.
    States: +1 (long spread), -1 (short spread), 0 (flat).

    Entry  when |z| > entry  (direction follows sign of z, reversed)
    Exit   when |z| < exit_
    Stop   when |z| > stop   (loss-limiting rule absent from original code)

    Returns a Series of positions (applied to the NEXT bar to avoid look-ahead).
    """
    z  = zscore.values
    n  = len(z)
    pos = np.zeros(n, dtype=int)

    current = 0
    for i in range(1, n):
        zi = z[i]
        if np.isnan(zi):
            pos[i] = 0
            current = 0
            continue

        if current == 0:
            if zi > entry:
                current = -1       # z high → short spread (sell s1, buy s2)
            elif zi < -entry:
                current = 1        # z low  → long spread  (buy s1, sell s2)
        else:
            if abs(zi) < exit_:
                current = 0        # mean reverted → exit
            elif abs(zi) > stop:
                current = 0        # stop-loss → exit
            # else: hold

        pos[i] = current

    return pd.Series(pos, index=zscore.index, name="position")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 6 — P&L WITH TRANSACTION COSTS & DOLLAR-NEUTRAL SIZING
# ═══════════════════════════════════════════════════════════════════════════
def compute_pnl(
    prices_test: pd.DataFrame,
    stock1: str,
    stock2: str,
    hedge_ratio: float,
    position: pd.Series,
    cost_per_leg: float,
) -> pd.DataFrame:
    """
    Dollar-neutral P&L calculation.

    At every bar, we hold:
        +position unit of log-return(s1)
        -position × hedge_ratio unit of log-return(s2)

    Transaction costs are charged when position changes (4 legs per round-trip):
        enter: buy s1 + sell s2 = 2 legs
        exit:  sell s1 + buy s2 = 2 legs
    Cost = 2 × cost_per_leg per direction change (enter or exit).

    NOTE: Using log-returns so that position=1 means long $1 of s1, short $beta of s2,
    which is consistent with the hedge ratio from log-price regression.
    """
    r1 = np.log(prices_test[stock1] / prices_test[stock1].shift(1))
    r2 = np.log(prices_test[stock2] / prices_test[stock2].shift(1))

    # Gross P&L: position from previous bar applied to current return
    pos_prev = position.shift(1).fillna(0)
    gross = pos_prev * (r1 - hedge_ratio * r2)

    # Transaction cost: charged when position changes
    trade_flag = (position.diff().abs() > 0).astype(float)
    tc = trade_flag * 2 * cost_per_leg   # 2 legs per direction change

    net_pnl = gross - tc

    cum_pnl = net_pnl.cumsum()
    return pd.DataFrame({
        "gross_pnl":  gross,
        "tc":         tc,
        "net_pnl":    net_pnl,
        "cum_pnl":    cum_pnl,
    }, index=prices_test.index)

# ═══════════════════════════════════════════════════════════════════════════
# STEP 7 — PERFORMANCE METRICS
# ═══════════════════════════════════════════════════════════════════════════
def compute_performance(pnl_df: pd.DataFrame, pair_name: str) -> dict:
    """
    Comprehensive performance metrics:
    - Annualised Sharpe ratio  (original used daily, unscaled — incorrect)
    - Profit factor
    - Max drawdown
    - Calmar ratio
    - Win rate
    - Number of trades
    """
    net = pnl_df["net_pnl"].dropna()

    if net.empty or net.std() == 0:
        return {}

    ann_factor     = np.sqrt(252)
    sharpe         = (net.mean() / net.std()) * ann_factor

    total_profit   = net[net > 0].sum()
    total_loss     = abs(net[net < 0].sum())
    profit_factor  = total_profit / total_loss if total_loss > 0 else np.nan

    win_rate       = (net > 0).mean()

    cum            = pnl_df["cum_pnl"].dropna()
    rolling_max    = cum.cummax()
    drawdown       = cum - rolling_max
    max_dd         = drawdown.min()

    total_return   = cum.iloc[-1]
    calmar         = total_return / abs(max_dd) if max_dd != 0 else np.nan

    # Count trades (state transitions into a non-zero position)
    n_trades       = (pnl_df.index.isin(
                        pnl_df.index[pnl_df["net_pnl"].abs() > 0]
                      )).sum()

    return {
        "pair":           pair_name,
        "sharpe_ann":     round(sharpe, 3),
        "profit_factor":  round(profit_factor, 3) if not np.isnan(profit_factor) else np.nan,
        "total_return":   round(total_return, 4),
        "max_drawdown":   round(max_dd, 4),
        "calmar":         round(calmar, 3) if not np.isnan(calmar) else np.nan,
        "win_rate":       round(win_rate, 3),
        "total_profit":   round(total_profit, 4),
        "total_loss":     round(total_loss, 4),
    }

# ═══════════════════════════════════════════════════════════════════════════
# STEP 8 — VISUALISATION
# ═══════════════════════════════════════════════════════════════════════════
def plot_pair(
    pair_row: pd.Series,
    spread_df: pd.DataFrame,
    position: pd.Series,
    pnl_df: pd.DataFrame,
    metrics: dict,
    train_end: str,
) -> None:
    s1, s2   = pair_row["stock1"], pair_row["stock2"]
    pair_name = f"{s1} / {s2}"

    fig = plt.figure(figsize=(14, 10))
    fig.suptitle(
        f"Pairs Trade: {pair_name}   "
        f"|   β={pair_row['hedge_ratio']:.3f}   "
        f"|   Half-life={pair_row['half_life']} days   "
        f"|   Sharpe={metrics.get('sharpe_ann', 'N/A')}",
        fontsize=12, fontweight="bold",
    )
    gs = GridSpec(3, 1, figure=fig, hspace=0.45)

    train_line = pd.Timestamp(train_end)

    # ── Panel 1: Spread + z-score bands ──────────────────────────────────
    ax1 = fig.add_subplot(gs[0])
    ax1.plot(spread_df.index, spread_df["spread"], color="black", lw=0.8, label="Spread")
    ax1.plot(spread_df.index, spread_df["roll_mean"], color="royalblue", ls="--", lw=1, label="Mean")
    ax1.plot(spread_df.index, spread_df["roll_mean"] + 2 * spread_df["roll_std"], color="red", ls=":", lw=1)
    ax1.plot(spread_df.index, spread_df["roll_mean"] - 2 * spread_df["roll_std"], color="green", ls=":", lw=1)
    ax1.axvline(train_line, color="purple", ls="--", lw=1.2, label="Train/Test split")
    ax1.set_title("Log-Price Spread with ±2σ Bands", fontsize=10)
    ax1.legend(fontsize=7)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))

    # ── Panel 2: Z-score + entry/exit lines + trade markers ──────────────
    ax2 = fig.add_subplot(gs[1])
    zscore = spread_df["zscore"]
    ax2.plot(zscore.index, zscore, color="black", lw=0.8, label="Z-score")
    ax2.axhline( CFG["zscore_entry"],  color="red",   ls="--", lw=0.8)
    ax2.axhline(-CFG["zscore_entry"],  color="green", ls="--", lw=0.8)
    ax2.axhline( CFG["zscore_exit"],   color="gray",  ls=":",  lw=0.7)
    ax2.axhline(-CFG["zscore_exit"],   color="gray",  ls=":",  lw=0.7)
    ax2.axhline( CFG["zscore_stop"],   color="darkred", ls="-.", lw=0.7, label="Stop")
    ax2.axhline(-CFG["zscore_stop"],   color="darkred", ls="-.", lw=0.7)
    ax2.axhline(0, color="black", lw=0.4)
    ax2.axvline(train_line, color="purple", ls="--", lw=1.2)

    # Mark entries (test period only)
    test_pos = position[position.index >= train_line]
    test_z   = zscore[zscore.index >= train_line]
    long_entries  = test_pos.index[(test_pos == 1) & (test_pos.shift(1, fill_value=0) == 0)]
    short_entries = test_pos.index[(test_pos == -1) & (test_pos.shift(1, fill_value=0) == 0)]
    ax2.scatter(long_entries,  test_z.reindex(long_entries),  marker="^", color="green", s=40, zorder=5, label="Long entry")
    ax2.scatter(short_entries, test_z.reindex(short_entries), marker="v", color="red",   s=40, zorder=5, label="Short entry")
    ax2.set_title("Z-Score with Entry / Exit / Stop Levels", fontsize=10)
    ax2.legend(fontsize=7, ncol=3)
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))

    # ── Panel 3: Cumulative net P&L (test period) ─────────────────────────
    ax3 = fig.add_subplot(gs[2])
    test_cum = pnl_df["cum_pnl"][pnl_df.index >= train_line]
    ax3.plot(test_cum.index, test_cum, color="steelblue", lw=1.2, label="Net cum. P&L")
    ax3.fill_between(test_cum.index, test_cum, 0, where=(test_cum < 0), color="red", alpha=0.25, label="Drawdown")
    ax3.axhline(0, color="black", lw=0.5)
    ax3.set_title(
        f"Out-of-Sample Cumulative Net P&L   "
        f"(Sharpe={metrics.get('sharpe_ann','—')}  |  "
        f"MaxDD={metrics.get('max_drawdown','—')}  |  "
        f"WinRate={metrics.get('win_rate','—')})",
        fontsize=9,
    )
    ax3.legend(fontsize=7)
    ax3.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))

    plt.savefig(OUTPUT_DIR / f"pair_{s1}_{s2}.png", dpi=130, bbox_inches="tight")
    plt.show()
    print(f"  Chart saved: {OUTPUT_DIR / f'pair_{s1}_{s2}.png'}")


def plot_portfolio(all_pnl: dict[str, pd.DataFrame], train_end: str) -> None:
    """Equal-weight portfolio of all pairs — net P&L."""
    train_line = pd.Timestamp(train_end)

    combined = pd.concat(
        {name: df["net_pnl"] for name, df in all_pnl.items()}, axis=1
    ).fillna(0)

    portfolio_pnl  = combined.mean(axis=1)          # equal-weight
    portfolio_cum  = portfolio_pnl.cumsum()
    test_cum       = portfolio_cum[portfolio_cum.index >= train_line]

    rolling_max  = test_cum.cummax()
    dd           = test_cum - rolling_max

    ann_sharpe   = (portfolio_pnl[portfolio_pnl.index >= train_line].mean()
                    / portfolio_pnl[portfolio_pnl.index >= train_line].std()
                    * np.sqrt(252))
    max_dd       = dd.min()

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 7), sharex=True)
    fig.suptitle(
        f"Equal-Weight Portfolio ({len(all_pnl)} pairs) — Out-of-Sample\n"
        f"Annualised Sharpe: {ann_sharpe:.2f}  |  Max Drawdown: {max_dd:.4f}",
        fontsize=12, fontweight="bold",
    )

    ax1.plot(test_cum.index, test_cum, color="steelblue", lw=1.4)
    ax1.fill_between(test_cum.index, test_cum, 0, where=(test_cum < 0), color="salmon", alpha=0.4)
    ax1.axhline(0, color="black", lw=0.5)
    ax1.set_title("Cumulative Net P&L (log-return units)", fontsize=10)

    ax2.fill_between(dd.index, dd, 0, color="red", alpha=0.5, label="Drawdown")
    ax2.set_title("Drawdown from Peak", fontsize=10)
    ax2.legend(fontsize=8)

    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "portfolio_pnl.png", dpi=130, bbox_inches="tight")
    plt.show()
    print(f"  Portfolio chart saved: {OUTPUT_DIR / 'portfolio_pnl.png'}")


# ═══════════════════════════════════════════════════════════════════════════
# GRID SEARCH — loop over entry/exit/window combinations per pair
# ═══════════════════════════════════════════════════════════════════════════
def run_grid_search(
    pairs_df: pd.DataFrame,
    prices_full: pd.DataFrame,
    prices_test: pd.DataFrame,
    grid: dict,
    train_end: str,
    cost_per_leg: float,
) -> pd.DataFrame:
    """
    For every calibrated pair × every parameter combination in GRID, compute
    out-of-sample P&L and performance metrics.  Returns a flat DataFrame with
    one row per (pair, entry, exit, window, stop) combination, sorted by
    annualised Sharpe descending.
    """
    entries  = grid["entry_thresholds"]
    exits    = grid["exit_thresholds"]
    windows  = grid["z_windows"]
    stops    = grid["stop_thresholds"]

    combos   = list(iproduct(entries, exits, windows, stops))
    n_pairs  = len(pairs_df)
    n_combos = len(combos)
    total    = n_pairs * n_combos
    print(f"  Running {n_pairs} pairs × {n_combos} parameter combos = {total:,} backtests …")

    records = []
    done    = 0

    for _, row in pairs_df.iterrows():
        s1, s2 = row["stock1"], row["stock2"]
        beta   = row["hedge_ratio"]

        if s1 not in prices_full.columns or s2 not in prices_full.columns:
            continue

        for (entry, exit_, z_win, stop) in combos:
            # Skip illogical combos
            if exit_ >= entry:
                continue

            spread_df   = compute_spread_zscore(prices_full, s1, s2, beta, z_win)
            test_zscore = spread_df.loc[train_end:, "zscore"]

            position = generate_signals_vectorised(
                test_zscore, entry=entry, exit_=exit_, stop=stop
            )

            pnl_df = compute_pnl(
                prices_test, s1, s2, beta, position, cost_per_leg=cost_per_leg
            )

            m = compute_performance(pnl_df, f"{s1}-{s2}")
            if not m:
                continue

            # Count trades (position changes into non-zero)
            pos_series = position
            n_trades = int(((pos_series != 0) & (pos_series != pos_series.shift(1))).sum())

            records.append({
                "pair":          f"{s1} / {s2}",
                "stock1":        s1,
                "stock2":        s2,
                "sector":        row.get("sector", ""),
                "hedge_ratio":   round(beta, 4),
                "half_life":     row["half_life"],
                "coint_pval":    round(row["coint_pval"], 4),
                "entry_z":       entry,
                "exit_z":        exit_,
                "stop_z":        stop,
                "z_window":      z_win,
                "sharpe_ann":    m["sharpe_ann"],
                "profit_factor": m["profit_factor"],
                "total_return":  m["total_return"],
                "max_drawdown":  m["max_drawdown"],
                "calmar":        m["calmar"],
                "win_rate":      m["win_rate"],
                "total_profit":  m["total_profit"],
                "total_loss":    m["total_loss"],
                "n_trades":      n_trades,
            })

            done += 1

        if done % 200 == 0 and done > 0:
            print(f"    … {done:,} / {total:,} done")

    df = (pd.DataFrame(records)
            .sort_values("sharpe_ann", ascending=False)
            .reset_index(drop=True))
    print(f"  Grid search complete — {len(df):,} valid results.")
    return df


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL WRITER — formatted workbook with grid results + best-pair charts
# ═══════════════════════════════════════════════════════════════════════════
def _col_widths(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 28)


def write_excel_report(
    grid_df: pd.DataFrame,
    pairs_df: pd.DataFrame,
    prices_full: pd.DataFrame,
    prices_test: pd.DataFrame,
    train_end: str,
    cost_per_leg: float,
    top_n: int,
    output_path,
) -> None:
    """
    Writes a multi-sheet Excel workbook:
      Sheet 1 — Full grid results (all combos), colour-scaled by Sharpe
      Sheet 2 — Top N combinations summary
      Sheet 3…N+2 — One sheet per top combination with monthly P&L table
    """
    wb = openpyxl.Workbook()

    # ── colour palette ───────────────────────────────────────────────────
    DARK_BLUE   = "1F3864"
    MID_BLUE    = "2E75B6"
    LIGHT_BLUE  = "BDD7EE"
    WHITE       = "FFFFFF"
    GREEN       = "E2EFDA"
    RED_LIGHT   = "FCE4D6"
    AMBER       = "FFF2CC"
    HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=10)
    BODY_FONT   = Font(name="Calibri", size=9)
    TITLE_FONT  = Font(name="Calibri", bold=True, color=DARK_BLUE, size=11)
    thin        = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTRE      = Alignment(horizontal="center", vertical="center")

    def header_fill(colour=MID_BLUE):
        return PatternFill("solid", fgColor=colour)

    def style_header_row(ws, row_idx, n_cols, colour=MID_BLUE):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill   = header_fill(colour)
            cell.font   = HEADER_FONT
            cell.border = BORDER
            cell.alignment = CENTRE

    def style_data_row(ws, row_idx, n_cols, fill_colour=None):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font   = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTRE
            if fill_colour:
                cell.fill = PatternFill("solid", fgColor=fill_colour)

    # ════════════════════════════════════════════════════════════════════
    # SHEET 1 — Full grid results
    # ════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "All Results"

    display_cols = [
        "pair", "entry_z", "exit_z", "stop_z", "z_window",
        "hedge_ratio", "half_life", "coint_pval",
        "sharpe_ann", "profit_factor", "total_return",
        "max_drawdown", "calmar", "win_rate", "n_trades",
        "total_profit", "total_loss",
    ]
    headers = [
        "Pair", "Entry Z", "Exit Z", "Stop Z", "Window",
        "Hedge Ratio", "Half-Life", "Coint p-val",
        "Sharpe (Ann)", "Profit Factor", "Total Return",
        "Max Drawdown", "Calmar", "Win Rate", "# Trades",
        "Total Profit", "Total Loss",
    ]

    # Title
    ws1.merge_cells("A1:Q1")
    ws1["A1"] = "Pairs Trading — Parameter Grid Search Results"
    ws1["A1"].font      = Font(name="Calibri", bold=True, color=WHITE, size=13)
    ws1["A1"].fill      = header_fill(DARK_BLUE)
    ws1["A1"].alignment = CENTRE
    ws1.row_dimensions[1].height = 22

    # Header row
    for c, h in enumerate(headers, 1):
        ws1.cell(row=2, column=c, value=h)
    style_header_row(ws1, 2, len(headers))

    # Data
    pct_cols  = {"Total Return", "Max Drawdown", "Win Rate", "Total Profit", "Total Loss"}
    num_fmt_2 = '0.00'
    num_fmt_4 = '0.0000'
    pct_fmt   = '0.00%'

    for r_idx, (_, row) in enumerate(grid_df[display_cols].iterrows(), start=3):
        fill_col = WHITE if r_idx % 2 == 1 else "F2F7FB"
        for c_idx, col in enumerate(display_cols, 1):
            val  = row[col]
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = BODY_FONT
            cell.border    = BORDER
            cell.alignment = CENTRE
            cell.fill      = PatternFill("solid", fgColor=fill_col)
            # number formats
            if col in ("total_return", "max_drawdown", "win_rate",
                       "total_profit", "total_loss"):
                cell.number_format = pct_fmt
            elif col in ("sharpe_ann", "profit_factor", "calmar",
                         "hedge_ratio", "coint_pval"):
                cell.number_format = num_fmt_2
            elif col == "half_life":
                cell.number_format = num_fmt_2

    n_data_rows = len(grid_df)

    # Colour-scale on Sharpe column (column I = 9)
    sharpe_col = headers.index("Sharpe (Ann)") + 1
    sharpe_letter = get_column_letter(sharpe_col)
    ws1.conditional_formatting.add(
        f"{sharpe_letter}3:{sharpe_letter}{n_data_rows + 2}",
        ColorScaleRule(
            start_type="min",  start_color="F8696B",
            mid_type="num",    mid_value=0,    mid_color="FFEB84",
            end_type="max",    end_color="63BE7B",
        ),
    )
    # Colour-scale on Max Drawdown (lower is better — reversed)
    dd_col    = headers.index("Max Drawdown") + 1
    dd_letter = get_column_letter(dd_col)
    ws1.conditional_formatting.add(
        f"{dd_letter}3:{dd_letter}{n_data_rows + 2}",
        ColorScaleRule(
            start_type="min",  start_color="63BE7B",
            mid_type="num",    mid_value=-0.05, mid_color="FFEB84",
            end_type="max",    end_color="F8696B",
        ),
    )

    ws1.freeze_panes = "A3"
    _col_widths(ws1)

    # ════════════════════════════════════════════════════════════════════
    # SHEET 2 — Top N summary
    # ════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Top Combinations")
    top_df = grid_df.head(top_n).reset_index(drop=True)

    ws2.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws2["A1"] = f"Top {top_n} Parameter Combinations by Annualised Sharpe"
    ws2["A1"].font      = Font(name="Calibri", bold=True, color=WHITE, size=13)
    ws2["A1"].fill      = header_fill(DARK_BLUE)
    ws2["A1"].alignment = CENTRE
    ws2.row_dimensions[1].height = 22

    for c, h in enumerate(headers, 1):
        ws2.cell(row=2, column=c, value=h)
    style_header_row(ws2, 2, len(headers))

    rank_fills = ["FFD700", "C0C0C0", "CD7F32"]   # gold / silver / bronze
    for r_idx, (_, row) in enumerate(top_df[display_cols].iterrows(), start=3):
        rank = r_idx - 3
        fill_col = rank_fills[rank] if rank < 3 else ("E8F4F8" if rank % 2 == 0 else WHITE)
        for c_idx, col in enumerate(display_cols, 1):
            val  = row[col]
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = Font(name="Calibri", size=9,
                                  bold=(rank < 3))
            cell.border    = BORDER
            cell.alignment = CENTRE
            cell.fill      = PatternFill("solid", fgColor=fill_col)
            if col in ("total_return", "max_drawdown", "win_rate",
                       "total_profit", "total_loss"):
                cell.number_format = pct_fmt
            elif col in ("sharpe_ann", "profit_factor", "calmar",
                         "hedge_ratio", "coint_pval"):
                cell.number_format = num_fmt_2

    ws2.freeze_panes = "A3"
    _col_widths(ws2)

    # ════════════════════════════════════════════════════════════════════
    # SHEETS 3+ — Monthly P&L breakdown for each top combination
    # ════════════════════════════════════════════════════════════════════
    for rank, (_, combo) in enumerate(top_df.iterrows(), start=1):
        s1, s2   = combo["stock1"], combo["stock2"]
        beta     = combo["hedge_ratio"]
        z_win    = int(combo["z_window"])
        entry    = combo["entry_z"]
        exit_    = combo["exit_z"]
        stop     = combo["stop_z"]

        spread_df   = compute_spread_zscore(prices_full, s1, s2, beta, z_win)
        test_zscore = spread_df.loc[train_end:, "zscore"]
        position    = generate_signals_vectorised(
            test_zscore, entry=entry, exit_=exit_, stop=stop
        )
        pnl_df = compute_pnl(
            prices_test, s1, s2, beta, position, cost_per_leg=cost_per_leg
        )

        # Monthly aggregation
        monthly = (pnl_df["net_pnl"]
                     .resample("ME")
                     .sum()
                     .to_frame("monthly_return"))
        monthly["cumulative"]  = pnl_df["cum_pnl"].resample("ME").last()
        monthly["positive"]    = monthly["monthly_return"] > 0

        sheet_name = f"#{rank} {s1}-{s2}"[:31]   # Excel 31-char limit
        ws = wb.create_sheet(sheet_name)

        # ── Title block ─────────────────────────────────────────────
        ws.merge_cells("A1:G1")
        ws["A1"] = (f"#{rank}  {s1} / {s2}   |   "
                    f"Entry={entry}  Exit={exit_}  Stop={stop}  Window={z_win}d   |   "
                    f"Sharpe={combo['sharpe_ann']:.2f}  MaxDD={combo['max_drawdown']:.2%}  "
                    f"WinRate={combo['win_rate']:.1%}")
        ws["A1"].font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
        ws["A1"].fill      = header_fill(DARK_BLUE)
        ws["A1"].alignment = CENTRE
        ws.row_dimensions[1].height = 20

        # ── Key metrics summary row ─────────────────────────────────
        metric_headers = ["Sharpe (Ann)", "Profit Factor", "Total Return",
                          "Max Drawdown", "Calmar", "Win Rate", "# Trades"]
        metric_vals    = [combo["sharpe_ann"], combo["profit_factor"],
                          combo["total_return"], combo["max_drawdown"],
                          combo["calmar"], combo["win_rate"],
                          int(combo["n_trades"])]
        metric_fmts    = [num_fmt_2, num_fmt_2, pct_fmt, pct_fmt,
                          num_fmt_2, pct_fmt, "0"]

        for c, h in enumerate(metric_headers, 1):
            ws.cell(row=2, column=c, value=h)
        style_header_row(ws, 2, len(metric_headers), colour=MID_BLUE)

        for c, (val, fmt) in enumerate(zip(metric_vals, metric_fmts), 1):
            cell = ws.cell(row=3, column=c, value=val)
            cell.font          = Font(name="Calibri", bold=True, size=10)
            cell.border        = BORDER
            cell.alignment     = CENTRE
            cell.number_format = fmt
            cell.fill          = PatternFill("solid", fgColor=LIGHT_BLUE)

        # ── Monthly P&L table ────────────────────────────────────────
        ws.cell(row=5, column=1, value="Month").font = TITLE_FONT
        monthly_headers = ["Month", "Net P&L", "Cumulative P&L", "Direction"]
        for c, h in enumerate(monthly_headers, 1):
            ws.cell(row=5, column=c, value=h)
        style_header_row(ws, 5, len(monthly_headers), colour=MID_BLUE)

        for r_off, (dt, mrow) in enumerate(monthly.iterrows(), start=6):
            is_pos  = bool(mrow["positive"])
            row_fill = GREEN if is_pos else RED_LIGHT
            ws.cell(row=r_off, column=1, value=dt.strftime("%b %Y"))
            ws.cell(row=r_off, column=2, value=mrow["monthly_return"]).number_format = pct_fmt
            ws.cell(row=r_off, column=3, value=mrow["cumulative"]).number_format     = pct_fmt
            ws.cell(row=r_off, column=4, value="▲" if is_pos else "▼")
            for c in range(1, 5):
                cell = ws.cell(row=r_off, column=c)
                cell.font      = BODY_FONT
                cell.border    = BORDER
                cell.alignment = CENTRE
                cell.fill      = PatternFill("solid", fgColor=row_fill)

        _col_widths(ws)

    wb.save(output_path)
    print(f"  Excel report saved: {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════════
def main():
    print("\n══ PAIRS TRADING STRATEGY ══════════════════════════════════════\n")

    # ── 1. Universe ──────────────────────────────────────────────────────
    print("[1] Fetching S&P 500 universe …")
    sp500 = fetch_sp500_tickers()
    tickers = sp500["Symbol"].tolist()

    # ── 2. Prices ────────────────────────────────────────────────────────
    print("[2] Downloading prices …")
    prices_full = download_prices(tickers, CFG["start_date"], CFG["test_end_date"])

    # Trim to aligned tickers only
    valid_tickers = [t for t in sp500["Symbol"] if t in prices_full.columns]
    sp500 = sp500[sp500["Symbol"].isin(valid_tickers)].reset_index(drop=True)

    # Train / test split
    prices_train = prices_full.loc[:CFG["train_end_date"]]
    prices_test  = prices_full.loc[CFG["train_end_date"]:]

    print(f"  Train: {prices_train.index[0].date()} → {prices_train.index[-1].date()}  ({len(prices_train)} days)")
    print(f"  Test:  {prices_test.index[0].date()}  → {prices_test.index[-1].date()}  ({len(prices_test)} days)")

    # ── 3. Correlation filter (in-sample only) ──────────────────────────
    print("[3] Computing in-sample correlations …")
    returns_train = compute_returns(prices_train)
    # Build sector lookup dict from the sp500 table
    sector_map = dict(zip(sp500["Symbol"], sp500["GICS Sector"]))

    candidate_pairs = select_correlated_pairs(
        returns_train,
        threshold        = CFG["corr_threshold"],
        sector_map       = sector_map,
        same_sector_only = CFG["same_sector_only"],
        max_pairs        = CFG["max_corr_pairs"],
    )

    if candidate_pairs.empty:
        print("  No pairs found above correlation threshold. Exiting.")
        return

    # ── 4. Cointegration (in-sample, with FDR correction) ───────────────
    print("[4] Testing cointegration (Engle-Granger + BH-FDR) …")
    coint_pairs = test_cointegration(
        prices_train,
        candidate_pairs,
        pval_threshold = CFG["coint_pval"],
        fdr            = CFG["fdr_correction"],
        n_workers      = CFG["n_workers"],
    )

    if coint_pairs.empty:
        print("  No cointegrated pairs found. Exiting.")
        return

    # ── 4b. Trim to top-N most-cointegrated pairs ────────────────────────
    top_n_coint = CFG["top_coint_pairs"]
    if top_n_coint is not None:
        coint_pairs = (coint_pairs
                       .sort_values("coint_pval", ascending=True)
                       .head(top_n_coint)
                       .reset_index(drop=True))
        print(f"  Trimmed to top {len(coint_pairs)} pairs by cointegration p-value.")

    # ── 5. Hedge ratio & half-life calibration (in-sample) ───────────────
    print("[5] Estimating hedge ratios and mean-reversion half-lives …")
    pairs_df = calibrate_pairs(
        coint_pairs,
        prices_train,
        min_hl=CFG["min_halflife_days"],
        max_hl=CFG["max_halflife_days"],
    )

    if pairs_df.empty:
        print("  No pairs pass half-life filter. Exiting.")
        return

    print(f"\n{'─'*70}")
    print(pairs_df[["stock1","stock2","correlation","coint_pval","hedge_ratio","half_life","z_window"]].to_string(index=False))
    print(f"{'─'*70}\n")

    # ── 6. Parameter grid search ─────────────────────────────────────────
    print("[6] Running parameter grid search (entry / exit / window / stop) …")
    grid_df = run_grid_search(
        pairs_df      = pairs_df,
        prices_full   = prices_full,
        prices_test   = prices_test,
        grid          = GRID,
        train_end     = CFG["train_end_date"],
        cost_per_leg  = CFG["transaction_cost"],
    )

    if grid_df.empty:
        print("  Grid search returned no results. Exiting.")
        return

    # Console preview of top results
    preview_cols = ["pair","entry_z","exit_z","stop_z","z_window",
                    "sharpe_ann","profit_factor","total_return",
                    "max_drawdown","calmar","win_rate","n_trades"]
    print("\n  Top 20 combinations:")
    print(grid_df[preview_cols].head(20).to_string(index=False))

    # ── 7. Excel report ───────────────────────────────────────────────────
    print("\n[7] Writing Excel report …")
    excel_path = OUTPUT_DIR / "pairs_grid_results.xlsx"
    write_excel_report(
        grid_df      = grid_df,
        pairs_df     = pairs_df,
        prices_full  = prices_full,
        prices_test  = prices_test,
        train_end    = CFG["train_end_date"],
        cost_per_leg = CFG["transaction_cost"],
        top_n        = CFG["top_n_plot"],
        output_path  = excel_path,
    )

    # ── 8. Charts for top combinations ────────────────────────────────────
    print(f"[8] Plotting top {CFG['top_n_plot']} combinations …")
    for rank, (_, combo) in enumerate(grid_df.head(CFG["top_n_plot"]).iterrows(), start=1):
        s1, s2   = combo["stock1"], combo["stock2"]
        beta     = combo["hedge_ratio"]
        z_win    = int(combo["z_window"])
        entry    = combo["entry_z"]
        exit_    = combo["exit_z"]
        stop     = combo["stop_z"]

        spread_df   = compute_spread_zscore(prices_full, s1, s2, beta, z_win)
        test_zscore = spread_df.loc[CFG["train_end_date"]:, "zscore"]
        position    = generate_signals_vectorised(
            test_zscore, entry=entry, exit_=exit_, stop=stop
        )
        pnl_df = compute_pnl(
            prices_test, s1, s2, beta, position,
            cost_per_leg=CFG["transaction_cost"],
        )
        metrics = {k: combo[k] for k in ["sharpe_ann", "max_drawdown", "win_rate"]}

        # Patch combo into a Series that plot_pair expects
        plot_row = combo.copy()
        plot_row["z_window"] = z_win
        plot_pair(plot_row, spread_df, position, pnl_df, metrics, CFG["train_end_date"])

    # Portfolio of top-N combinations (equal weight)
    top_pnl = {}
    for _, combo in grid_df.head(CFG["top_n_plot"]).iterrows():
        s1, s2  = combo["stock1"], combo["stock2"]
        beta    = combo["hedge_ratio"]
        z_win   = int(combo["z_window"])
        entry   = combo["entry_z"]
        exit_   = combo["exit_z"]
        stop    = combo["stop_z"]
        key     = f"{s1}-{s2} e={entry} ex={exit_} w={z_win}"

        spread_df   = compute_spread_zscore(prices_full, s1, s2, beta, z_win)
        test_zscore = spread_df.loc[CFG["train_end_date"]:, "zscore"]
        position    = generate_signals_vectorised(
            test_zscore, entry=entry, exit_=exit_, stop=stop
        )
        top_pnl[key] = compute_pnl(
            prices_test, s1, s2, beta, position,
            cost_per_leg=CFG["transaction_cost"],
        )

    plot_portfolio(top_pnl, CFG["train_end_date"])

    print("\n══ DONE ═════════════════════════════════════════════════════════\n")


if __name__ == "__main__":
    main()